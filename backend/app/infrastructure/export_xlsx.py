"""Geração server-side dos relatórios em XLSX (formatado, A4) e CSV.

Padrão de marca FieldTech: cabeçalhos FieldBlue (#00AEEF) com texto branco,
linhas zebradas em #F4F7FA, texto #0D1B2A. Planilha pronta para impressão A4.
"""
import csv
import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.application.dtos import MonthlyRecord, MonthlyReport, MonthlySummary

MONTH_NAMES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]
WEEKDAY_ABBR = ["seg", "ter", "qua", "qui", "sex", "sáb", "dom"]

DOC_CODE = "FPRESI-RH-0134 Rev.04"

FIELD_BLUE = "FF00AEEF"
FIELD_DARK = "FF0D1B2A"
BG_SUBTLE = "FFF4F7FA"
BG_TOTAL = "FFDDEFF9"
BORDER_CLR = "FFD5DBE0"
STATUS_PEND = "FFFFF3D6"
STATUS_REJ = "FFF8D7DA"

_thin = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEAD_FILL = PatternFill("solid", fgColor=FIELD_BLUE)
SUBTLE_FILL = PatternFill("solid", fgColor=BG_SUBTLE)
TOTAL_FILL = PatternFill("solid", fgColor=BG_TOTAL)
HEAD_FONT = Font(name="Aptos", bold=True, color="FFFFFFFF", size=10)
CELL_FONT = Font(name="Aptos", size=10, color=FIELD_DARK)
BOLD_FONT = Font(name="Aptos", size=10, bold=True, color=FIELD_DARK)
TITLE_FONT = Font(name="Aptos Display", bold=True, size=15, color=FIELD_BLUE)
SUB_FONT = Font(name="Aptos", size=9, color="FF5A6B7A")

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

ABONO_LABELS = {"AB": "Abono", "AT": "Atestado", "VG": "Viagem", "FA": "Falta", "FE": "Folga"}
STATUS_LABELS = {"aprovado": "Aprovado", "pendente": "Pendente", "reprovado": "Reprovado"}


def _hhmm(minutes) -> str:
    if minutes is None:
        return ""
    m = int(minutes)
    return f"{m // 60}:{m % 60:02d}"


def _signed(minutes) -> str:
    if minutes is None:
        return ""
    m = int(minutes)
    sign = "-" if m < 0 else "+"
    m = abs(m)
    return f"{sign}{m // 60}:{m % 60:02d}"


def _br_date(iso: str) -> str:
    d = date.fromisoformat(iso)
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def _safe_sheet_name(name: str, used: set[str]) -> str:
    bad = set('[]:*?/\\')
    clean = "".join("-" if c in bad else c for c in name)[:28].strip() or "Colab"
    candidate = clean
    i = 2
    while candidate.lower() in used:
        candidate = f"{clean[:25]}_{i}"
        i += 1
    used.add(candidate.lower())
    return candidate


def _apply_print(ws, landscape: bool, cols: int) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 24


RESUMO_HEAD = [
    "Colaborador", "Úteis", "Sáb", "Dom/Fer", "Referência", "Trabalhado",
    "Saldo", "Extra 50%", "Extra 100%", "Adic. Not.", "Faltas", "Atest.",
    "Folgas", "Pend.",
]
RESUMO_WIDTHS = [26, 7, 6, 8, 12, 12, 11, 11, 11, 11, 8, 8, 8, 7]


def _build_resumo(ws, rep: MonthlyReport) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RESUMO_HEAD))
    t = ws.cell(row=1, column=1, value="FIELD TECHNOLOGY — Banco de Horas / Folha de Ponto")
    t.font = TITLE_FONT
    t.alignment = LEFT
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(RESUMO_HEAD))
    sub = ws.cell(
        row=2, column=1,
        value=f"Referência: {MONTH_NAMES[rep.month - 1]}/{rep.year}    ·    "
              f"Doc: {DOC_CODE}    ·    Gerado em {date.today().strftime('%d/%m/%Y')}",
    )
    sub.font = SUB_FONT
    sub.alignment = LEFT

    head_row = 4
    for i, h in enumerate(RESUMO_HEAD, start=1):
        ws.cell(row=head_row, column=i, value=h)
    _style_header_row(ws, head_row, len(RESUMO_HEAD))

    r = head_row + 1
    for idx, s in enumerate(rep.summary):
        vals = [
            s.employee_name, s.days_h1, s.days_h2, s.days_h3,
            _hhmm(s.reference_minutes), _hhmm(s.worked_minutes), _signed(s.balance),
            _hhmm(s.extra50_minutes), _hhmm(s.extra100_minutes), _hhmm(s.night_bonus_minutes),
            s.faltas, s.atestados, s.folgas, s.pending,
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else (RIGHT if c >= 5 else CENTER)
            if idx % 2 == 1:
                cell.fill = SUBTLE_FILL
        r += 1

    tot_worked = sum(s.worked_minutes for s in rep.summary)
    tot_ref = sum(s.reference_minutes for s in rep.summary)
    tot_bal = sum(s.balance for s in rep.summary)
    tot_e100 = sum(s.extra100_minutes for s in rep.summary)
    tot_night = sum(s.night_bonus_minutes for s in rep.summary)
    total_vals = [
        "TOTAL",
        sum(s.days_h1 for s in rep.summary), sum(s.days_h2 for s in rep.summary),
        sum(s.days_h3 for s in rep.summary),
        _hhmm(tot_ref), _hhmm(tot_worked), _signed(tot_bal),
        _hhmm(max(tot_bal, 0)), _hhmm(tot_e100), _hhmm(tot_night),
        sum(s.faltas for s in rep.summary), sum(s.atestados for s in rep.summary),
        sum(s.folgas for s in rep.summary), rep.pending_records,
    ]
    for c, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = BOLD_FONT
        cell.border = BORDER
        cell.fill = TOTAL_FILL
        cell.alignment = LEFT if c == 1 else (RIGHT if c >= 5 else CENTER)

    for i, w in enumerate(RESUMO_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    _apply_print(ws, landscape=True, cols=len(RESUMO_HEAD))


DETAIL_HEAD = [
    "Data", "Dia", "Tipo", "Entrada", "Início Int.", "Fim Int.", "Saída",
    "Trab.", "Ref.", "Saldo", "E 50%", "E 100%", "Noturno", "Abono", "Status", "Obs.",
]
DETAIL_WIDTHS = [11, 5, 5, 8, 9, 9, 8, 8, 8, 9, 8, 8, 8, 9, 10, 24]


def _period_days(year: int, month: int, start: str | None, end: str | None) -> list[str]:
    """Todas as datas do mês (ou do intervalo filtrado), em ordem."""
    from calendar import monthrange
    last = monthrange(year, month)[1]
    ini = date(year, month, 1)
    fim = date(year, month, last)
    if start:
        try: ini = max(ini, date.fromisoformat(start))
        except ValueError: pass
    if end:
        try: fim = min(fim, date.fromisoformat(end))
        except ValueError: pass
    out, d = [], ini
    while d <= fim:
        out.append(d.isoformat())
        d += timedelta(days=1)
    return out


def _build_employee_sheet(ws, s: MonthlySummary, records: list[MonthlyRecord],
                          year: int = 0, month: int = 0, h1: int = 480, h2: int = 240,
                          schedule: tuple | None = None,
                          start: str | None = None, end: str | None = None) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DETAIL_HEAD))
    t = ws.cell(row=1, column=1, value=f"Folha de Ponto — {s.employee_name}")
    t.font = TITLE_FONT
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(DETAIL_HEAD))
    sub = ws.cell(
        row=2, column=1,
        value=(f"Saldo do mês: {_signed(s.balance)}   ·   Trabalhado: {_hhmm(s.worked_minutes)}   ·   "
               f"Referência: {_hhmm(s.reference_minutes)}   ·   Dias: {s.days} "
               f"(H1 {s.days_h1} / H2 {s.days_h2} / H3 {s.days_h3})"),
    )
    sub.font = SUB_FONT

    head_row = 4
    for i, h in enumerate(DETAIL_HEAD, start=1):
        ws.cell(row=head_row, column=i, value=h)
    _style_header_row(ws, head_row, len(DETAIL_HEAD))

    r = head_row + 1
    by_date = {rec.date: rec for rec in records}
    all_days = _period_days(year, month, start, end) if year and month else sorted(by_date)

    for idx, iso in enumerate(all_days):
        rec = by_date.get(iso)
        d = date.fromisoformat(iso)

        if rec is None:
            # Dia sem registro: mostra a referência esperada e a situação
            try:
                from app.domain import accounting as _acc
                ref = _acc.employee_reference(s.employee_id, iso, None, h1, h2,
                                              _acc.weekday_reference_override(schedule, iso),
                                              on_leave=_acc.is_on_leave(s.employee_id, iso))
                dtype = _acc.day_type(iso)
                on_leave = _acc.is_on_leave(s.employee_id, iso)
            except Exception:
                ref, dtype, on_leave = 0, "", False
            situacao = ("FÉRIAS/LICENÇA" if on_leave
                        else "SEM REGISTRO" if ref > 0 else "descanso")
            vals = [
                _br_date(iso), WEEKDAY_ABBR[d.weekday()], dtype,
                "", "", "", "",
                "", _hhmm(ref), _signed(-ref) if ref > 0 else "",
                "", "", "", "", situacao, "",
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = CELL_FONT
                cell.border = BORDER
                cell.alignment = LEFT if c in (1, 16) else (RIGHT if 8 <= c <= 13 else CENTER)
                if ref > 0 and not on_leave:
                    cell.fill = PatternFill("solid", fgColor="FDE2E2")   # falta
                elif idx % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="F7FAFF")
            r += 1
            continue

        vals = [
            _br_date(rec.date), WEEKDAY_ABBR[d.weekday()], rec.day_type or "",
            rec.entry_time or "", rec.break_start or "", rec.break_end or "", rec.exit_time or "",
            _hhmm(rec.worked_minutes), _hhmm(rec.standard_minutes), _signed(rec.overtime_minutes),
            _hhmm(rec.extra50_minutes), _hhmm(rec.extra100_minutes), _hhmm(rec.night_bonus_minutes),
            ABONO_LABELS.get(rec.abono_code, "") if rec.abono_code else "",
            STATUS_LABELS.get(rec.status, rec.status),
            rec.note or "",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c in (1, 16) else (RIGHT if 8 <= c <= 13 else CENTER)
            if rec.status == "pendente":
                cell.fill = PatternFill("solid", fgColor=STATUS_PEND)
            elif rec.status == "reprovado":
                cell.fill = PatternFill("solid", fgColor=STATUS_REJ)
            elif idx % 2 == 1:
                cell.fill = SUBTLE_FILL
        r += 1

    total_vals = [
        "TOTAL", "", "", "", "", "", "",
        _hhmm(s.worked_minutes), _hhmm(s.reference_minutes), _signed(s.balance),
        _hhmm(s.extra50_minutes), _hhmm(s.extra100_minutes), _hhmm(s.night_bonus_minutes),
        "", "", "",
    ]
    for c, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.alignment = LEFT if c == 1 else (RIGHT if 8 <= c <= 13 else CENTER)

    r += 2
    ws.cell(row=r, column=1,
            value="Legenda abono: AB=Abono · AT=Atestado · VG=Viagem · FA=Falta · FE=Folga  |  "
                  "Tipo: H1=dia útil (8h) · H2=sábado (4h) · H3=domingo/feriado").font = SUB_FONT

    for i, w in enumerate(DETAIL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    _apply_print(ws, landscape=True, cols=len(DETAIL_HEAD))


def build_monthly_xlsx(rep: MonthlyReport, *, h1: int = 480, h2: int = 240,
                       schedules: dict[int, tuple] | None = None,
                       start: str | None = None, end: str | None = None) -> bytes:
    """Planilha do mês. A aba de cada colaborador traz TODOS os dias do período
    (inclusive os sem registro, marcados como falta/descanso)."""
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Mensal"
    _build_resumo(ws_resumo, rep)

    by_emp: dict[int, list[MonthlyRecord]] = {}
    for rec in rep.records:
        by_emp.setdefault(rec.employee_id, []).append(rec)

    used_names: set[str] = {"resumo mensal"}
    for s in rep.summary:
        recs = sorted(by_emp.get(s.employee_id, []), key=lambda x: x.date)
        ws = wb.create_sheet(_safe_sheet_name(s.employee_name, used_names))
        _build_employee_sheet(ws, s, recs, rep.year, rep.month, h1, h2,
                              (schedules or {}).get(s.employee_id), start, end)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_monthly_csv(rep: MonthlyReport, *, h1: int = 480, h2: int = 240,
                      schedules: dict[int, tuple] | None = None,
                      start: str | None = None, end: str | None = None) -> str:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow([
        "Data", "Colaborador", "Dia", "Tipo", "Entrada", "Inicio_Intervalo",
        "Fim_Intervalo", "Saida", "Trabalhado", "Referencia", "Saldo",
        "Extra_50", "Extra_100", "Adicional_Noturno", "Abono", "Status",
        "Retroativo", "Observacao",
    ])
    # Linhas de TODOS os dias do período — inclusive sem registro
    dias = _period_days(rep.year, rep.month, start, end) if rep.year and rep.month else []
    por_emp: dict[int, dict[str, MonthlyRecord]] = {}
    for rec in rep.records:
        por_emp.setdefault(rec.employee_id, {})[rec.date] = rec

    linhas: list[MonthlyRecord | tuple] = []
    for s_ in sorted(rep.summary, key=lambda x: x.employee_name.lower()):
        recs = por_emp.get(s_.employee_id, {})
        for iso in (dias or sorted(recs)):
            r_ = recs.get(iso)
            if r_ is not None:
                linhas.append(r_)
            else:
                try:
                    from app.domain import accounting as _acc
                    sch = (schedules or {}).get(s_.employee_id)
                    ref = _acc.employee_reference(s_.employee_id, iso, None, h1, h2,
                                                  _acc.weekday_reference_override(sch, iso),
                                                  on_leave=_acc.is_on_leave(s_.employee_id, iso))
                    dtype = _acc.day_type(iso)
                    on_leave = _acc.is_on_leave(s_.employee_id, iso)
                except Exception:
                    ref, dtype, on_leave = 0, "", False
                d0 = date.fromisoformat(iso)
                w.writerow([
                    iso, s_.employee_name, WEEKDAY_ABBR[d0.weekday()], dtype,
                    "", "", "", "", "", _hhmm(ref), _signed(-ref) if ref > 0 else "",
                    "", "", "",
                    "", ("FERIAS/LICENCA" if on_leave else "SEM REGISTRO" if ref > 0 else "descanso"),
                    "", "",
                ])

    for rec in linhas:
        d = date.fromisoformat(rec.date)
        w.writerow([
            rec.date, rec.employee_name, WEEKDAY_ABBR[d.weekday()], rec.day_type or "",
            rec.entry_time or "", rec.break_start or "", rec.break_end or "", rec.exit_time or "",
            _hhmm(rec.worked_minutes), _hhmm(rec.standard_minutes), _signed(rec.overtime_minutes),
            _hhmm(rec.extra50_minutes), _hhmm(rec.extra100_minutes), _hhmm(rec.night_bonus_minutes),
            ABONO_LABELS.get(rec.abono_code, "") if rec.abono_code else "",
            STATUS_LABELS.get(rec.status, rec.status),
            "Sim" if rec.is_retroactive else "Não",
            (rec.note or "").replace("\n", " "),
        ])
    return buf.getvalue()
